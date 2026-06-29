"""
Data Cleaning & Reporting Automation Pipeline
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
import json
import re

# ─────────────────────────────────────────────
def load_data(filepath):
    return pd.read_csv(filepath).copy()

# ─────────────────────────────────────────────
def clean_data(df):
    log = []
    original_shape = df.shape
    df.columns = df.columns.str.strip().str.upper()

    df['REGION'] = df['REGION'].str.strip().str.title()
    log.append({"step":"Region Standardization","issue":"Mixed case regions","fixed":"Standardized to Title Case"})

    df['STATUS'] = df['STATUS'].str.strip().str.title()
    log.append({"step":"Status Standardization","issue":"Mixed case statuses","fixed":"Standardized to Title Case"})

    df['CATEGORY'] = df['CATEGORY'].str.strip().str.title()
    log.append({"step":"Category Standardization","issue":"Mixed case categories","fixed":"Standardized to Title Case"})

    def parse_date(val):
        val = str(val).strip()
        for fmt in ('%Y-%m-%d','%d-%m-%Y','%m/%d/%Y','%d/%m/%Y'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
        return None

    bad = df['ORDER_DATE'].apply(lambda x: len(str(x).split('-')[0])==2).sum()
    df['ORDER_DATE'] = df['ORDER_DATE'].apply(parse_date)
    log.append({"step":"Date Normalization","issue":f"{bad} dates in dd-mm-yyyy","fixed":"Converted all to yyyy-mm-dd"})

    missing_names = df['CUSTOMER_NAME'].isna().sum()
    df['CUSTOMER_NAME'] = df['CUSTOMER_NAME'].fillna('Unknown Customer')
    log.append({"step":"Missing Names","issue":f"{missing_names} empty names","fixed":"Filled with 'Unknown Customer'"})

    email_pat = r'^[\w\.-]+@[\w\.-]+\.\w{2,}$'
    invalid = ~df['EMAIL'].str.match(email_pat, na=False)
    inv_count = invalid.sum()
    df.loc[invalid,'EMAIL'] = 'invalid@unknown.com'
    log.append({"step":"Invalid Emails","issue":f"{inv_count} malformed emails","fixed":"Replaced with placeholder"})

    missing_qty = df['QUANTITY'].isna().sum()
    zero_qty = (df['QUANTITY']==0).sum()
    med = int(df['QUANTITY'].median())
    df['QUANTITY'] = df['QUANTITY'].fillna(med)
    df.loc[df['QUANTITY']==0,'QUANTITY'] = med
    df['QUANTITY'] = df['QUANTITY'].astype(int)
    log.append({"step":"Missing/Zero Quantity","issue":f"{missing_qty} missing, {zero_qty} zeros","fixed":f"Replaced with median ({med})"})

    dup = df.duplicated(subset=['CUSTOMER_NAME','EMAIL','PRODUCT','ORDER_DATE']).sum()
    df = df.drop_duplicates(subset=['CUSTOMER_NAME','EMAIL','PRODUCT','ORDER_DATE'], keep='first')
    log.append({"step":"Duplicate Removal","issue":f"{dup} duplicate rows","fixed":f"{dup} rows removed"})

    df['TOTAL_REVENUE'] = df['QUANTITY'] * df['UNIT_PRICE']
    df['MONTH'] = pd.to_datetime(df['ORDER_DATE'].astype(str)).dt.strftime('%B %Y')
    log.append({"step":"Derived Columns","issue":"No revenue/month columns","fixed":"Added TOTAL_REVENUE and MONTH"})

    final_shape = df.shape
    log.append({"step":"Summary","issue":f"Original: {original_shape[0]} rows × {original_shape[1]} cols","fixed":f"Cleaned: {final_shape[0]} rows × {final_shape[1]} cols"})

    return df.reset_index(drop=True), log

# ─────────────────────────────────────────────
def compute_summaries(df):
    return {
        'by_region': df.groupby('REGION').agg(
            Total_Orders=('ORDER_ID','count'),
            Total_Revenue=('TOTAL_REVENUE','sum'),
            Avg_Order_Value=('TOTAL_REVENUE','mean')
        ).reset_index(),
        'by_category': df.groupby('CATEGORY').agg(
            Total_Orders=('ORDER_ID','count'),
            Total_Revenue=('TOTAL_REVENUE','sum'),
            Total_Qty=('QUANTITY','sum')
        ).reset_index(),
        'by_product': df.groupby('PRODUCT').agg(
            Total_Revenue=('TOTAL_REVENUE','sum'),
            Units_Sold=('QUANTITY','sum'),
            Order_Count=('ORDER_ID','count')
        ).reset_index().sort_values('Total_Revenue',ascending=False),
        'by_status': df['STATUS'].value_counts().reset_index().rename(columns={'STATUS':'Status','count':'Count'}),
    }

# ─────────────────────────────────────────────
def write_excel_report(df_clean, log, summaries, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    # Palette
    DARK_BLUE   = "1B2A4A"
    MID_BLUE    = "2E5FA3"
    LIGHT_BLUE  = "EAF4FB"
    WHITE       = "FFFFFF"
    GREEN       = "27AE60"
    ORANGE      = "E67E22"

    def thin_border():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    def style_header(cell, text, bg=DARK_BLUE, fg=WHITE, sz=11):
        cell.value = text
        cell.font = Font(name='Calibri', bold=True, color=fg, size=sz)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def style_data(cell, value, fmt=None, bg=WHITE):
        cell.value = value
        cell.font = Font(name='Calibri', size=10)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()
        if fmt:
            cell.number_format = fmt

    # ── SHEET 1: Dashboard ──────────────────────────────────────────────
    ws1 = wb.create_sheet("Dashboard")
    ws1.sheet_view.showGridLines = False

    # Title
    ws1.row_dimensions[1].height = 8
    ws1.row_dimensions[2].height = 44
    ws1.row_dimensions[3].height = 20
    ws1.row_dimensions[4].height = 14

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    c = ws1.cell(2, 1)
    c.value = "  SALES DATA AUTOMATION — CLEANING & REPORTING DASHBOARD"
    c.font = Font(name='Calibri', bold=True, size=18, color=WHITE)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center')

    ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
    sub = ws1.cell(3, 1)
    sub.value = f"  Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}   |   Source: raw_sales_data.csv   |   Records after cleaning: {len(df_clean)}"
    sub.font = Font(name='Calibri', size=9, color=WHITE)
    sub.fill = PatternFill('solid', fgColor=MID_BLUE)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    # KPI cards (5 cards, columns 1-2, 3-4, 5-6, 7-8, 9-10)
    kpis = [
        ("Total Orders", len(df_clean), None, MID_BLUE),
        ("Total Revenue", f"\u20b9{int(df_clean['TOTAL_REVENUE'].sum()):,}", None, "1B6CA8"),
        ("Avg Order Value", f"\u20b9{int(df_clean['TOTAL_REVENUE'].mean()):,}", None, GREEN),
        ("Products", df_clean['PRODUCT'].nunique(), None, "8E44AD"),
        ("Regions", df_clean['REGION'].nunique(), None, ORANGE),
    ]
    ws1.row_dimensions[5].height = 8
    ws1.row_dimensions[6].height = 28
    ws1.row_dimensions[7].height = 34
    ws1.row_dimensions[8].height = 20
    ws1.row_dimensions[9].height = 8

    for i, (label, value, fmt, color) in enumerate(kpis):
        sc = i * 2 + 1
        ec = sc + 1
        ws1.merge_cells(start_row=6, start_column=sc, end_row=6, end_column=ec)
        ws1.merge_cells(start_row=7, start_column=sc, end_row=7, end_column=ec)
        ws1.merge_cells(start_row=8, start_column=sc, end_row=8, end_column=ec)
        lbl_cell = ws1.cell(6, sc)
        lbl_cell.value = label.upper()
        lbl_cell.font = Font(name='Calibri', size=8, color="CCDDEE", bold=True)
        lbl_cell.fill = PatternFill('solid', fgColor=color)
        lbl_cell.alignment = Alignment(horizontal='center', vertical='center')

        val_cell = ws1.cell(7, sc)
        val_cell.value = value
        val_cell.font = Font(name='Calibri', bold=True, size=16, color=WHITE)
        val_cell.fill = PatternFill('solid', fgColor=color)
        val_cell.alignment = Alignment(horizontal='center', vertical='center')

        pad_cell = ws1.cell(8, sc)
        pad_cell.fill = PatternFill('solid', fgColor=color)

    # KPI col widths
    for ci in range(1, 11):
        ws1.column_dimensions[get_column_letter(ci)].width = 14

    # Region summary mini-table on dashboard
    ws1.row_dimensions[11].height = 24
    reg = summaries['by_region']
    style_header(ws1.cell(11, 1), "REGION", bg=DARK_BLUE, sz=10)
    style_header(ws1.cell(11, 2), "ORDERS", bg=DARK_BLUE, sz=10)
    style_header(ws1.cell(11, 3), "REVENUE", bg=DARK_BLUE, sz=10)

    for ri, row in reg.iterrows():
        r = ri + 12
        ws1.row_dimensions[r].height = 20
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        style_data(ws1.cell(r, 1), row['REGION'], bg=bg)
        style_data(ws1.cell(r, 2), row['Total_Orders'], bg=bg)
        style_data(ws1.cell(r, 3), row['Total_Revenue'], fmt='\u20b9#,##0', bg=bg)

    # Product summary mini-table
    ws1.cell(11, 5).value = "PRODUCT"
    ws1.cell(11, 6).value = "REVENUE"
    for hc in [5,6]:
        style_header(ws1.cell(11, hc), ws1.cell(11, hc).value, bg=DARK_BLUE, sz=10)

    prod = summaries['by_product'].reset_index(drop=True)
    for ri, row in prod.iterrows():
        r = ri + 12
        ws1.row_dimensions[r].height = 20
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        style_data(ws1.cell(r, 5), row['PRODUCT'], bg=bg)
        style_data(ws1.cell(r, 6), row['Total_Revenue'], fmt='\u20b9#,##0', bg=bg)

    # ── SHEET 2: Cleaned Data ────────────────────────────────────────────
    ws2 = wb.create_sheet("Cleaned Data")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A2'

    ws2.row_dimensions[1].height = 26
    for ci, col in enumerate(df_clean.columns, 1):
        style_header(ws2.cell(1, ci), col.replace('_',' '), bg=DARK_BLUE)

    col_widths_2 = [9,18,22,10,14,14,10,12,14,12,16,14]
    for ci, w in enumerate(col_widths_2[:len(df_clean.columns)], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for ri, (_, row) in enumerate(df_clean.iterrows(), 2):
        ws2.row_dimensions[ri].height = 18
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        for ci, col in enumerate(df_clean.columns, 1):
            v = row[col]
            fmt = '\u20b9#,##0' if col == 'TOTAL_REVENUE' else None
            style_data(ws2.cell(ri, ci), v if not isinstance(v, (int,float)) else v, fmt=fmt, bg=bg)

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(df_clean.columns))}1"

    # ── SHEET 3: Cleaning Log ────────────────────────────────────────────
    ws3 = wb.create_sheet("Cleaning Log")
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 8
    ws3.row_dimensions[2].height = 36
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    hc = ws3.cell(2, 1)
    hc.value = "DATA CLEANING LOG"
    hc.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    hc.fill = PatternFill('solid', fgColor=DARK_BLUE)
    hc.alignment = Alignment(horizontal='center', vertical='center')

    ws3.row_dimensions[4].height = 24
    for ci, h in enumerate(['Cleaning Step','Issue Detected','Action Taken'], 1):
        style_header(ws3.cell(4, ci), h, bg=MID_BLUE)

    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 38
    ws3.column_dimensions['C'].width = 40

    for ri, entry in enumerate(log, 5):
        ws3.row_dimensions[ri].height = 22
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        for ci, key in enumerate(['step','issue','fixed'], 1):
            c = ws3.cell(ri, ci)
            c.value = entry[key]
            c.fill = PatternFill('solid', fgColor=bg)
            c.font = Font(name='Calibri', size=10)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border = thin_border()

    # ── SHEET 4: Region Summary ──────────────────────────────────────────
    ws4 = wb.create_sheet("By Region")
    ws4.sheet_view.showGridLines = False
    ws4.row_dimensions[1].height = 8
    ws4.row_dimensions[2].height = 36
    ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    h4 = ws4.cell(2, 1)
    h4.value = "REVENUE BY REGION"
    h4.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    h4.fill = PatternFill('solid', fgColor=DARK_BLUE)
    h4.alignment = Alignment(horizontal='center', vertical='center')

    ws4.row_dimensions[4].height = 24
    for ci, h in enumerate(['Region','Total Orders','Total Revenue','Avg Order Value'], 1):
        style_header(ws4.cell(4, ci), h, bg=MID_BLUE)
    for ci, w in enumerate([20,16,22,24], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in reg.iterrows():
        r = ri + 5
        ws4.row_dimensions[r].height = 20
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        style_data(ws4.cell(r, 1), row['REGION'], bg=bg)
        style_data(ws4.cell(r, 2), row['Total_Orders'], bg=bg)
        style_data(ws4.cell(r, 3), row['Total_Revenue'], fmt='\u20b9#,##0', bg=bg)
        style_data(ws4.cell(r, 4), row['Avg_Order_Value'], fmt='\u20b9#,##0', bg=bg)

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Revenue by Region"
    chart.y_axis.title = "Revenue"
    chart.style = 10
    chart.width = 18
    chart.height = 10
    n = len(reg)
    data_ref = Reference(ws4, min_col=3, min_row=4, max_row=4+n)
    cats_ref = Reference(ws4, min_col=1, min_row=5, max_row=4+n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws4.add_chart(chart, f"A{n+8}")

    # ── SHEET 5: Product Summary ─────────────────────────────────────────
    ws5 = wb.create_sheet("By Product")
    ws5.sheet_view.showGridLines = False
    ws5.row_dimensions[1].height = 8
    ws5.row_dimensions[2].height = 36
    ws5.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    h5 = ws5.cell(2, 1)
    h5.value = "PRODUCT PERFORMANCE"
    h5.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    h5.fill = PatternFill('solid', fgColor=DARK_BLUE)
    h5.alignment = Alignment(horizontal='center', vertical='center')

    ws5.row_dimensions[4].height = 24
    for ci, h in enumerate(['Product','Total Revenue','Units Sold','Order Count'], 1):
        style_header(ws5.cell(4, ci), h, bg=MID_BLUE)
    for ci, w in enumerate([20,22,14,14], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    prod5 = summaries['by_product'].reset_index(drop=True)
    for ri, row in prod5.iterrows():
        r = ri + 5
        ws5.row_dimensions[r].height = 20
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        style_data(ws5.cell(r, 1), row['PRODUCT'], bg=bg)
        style_data(ws5.cell(r, 2), row['Total_Revenue'], fmt='\u20b9#,##0', bg=bg)
        style_data(ws5.cell(r, 3), int(row['Units_Sold']), bg=bg)
        style_data(ws5.cell(r, 4), int(row['Order_Count']), bg=bg)

    # ── SHEET 6: Order Status ────────────────────────────────────────────
    ws6 = wb.create_sheet("Order Status")
    ws6.sheet_view.showGridLines = False
    ws6.row_dimensions[1].height = 8
    ws6.row_dimensions[2].height = 36
    ws6.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    h6 = ws6.cell(2, 1)
    h6.value = "ORDER STATUS BREAKDOWN"
    h6.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    h6.fill = PatternFill('solid', fgColor=DARK_BLUE)
    h6.alignment = Alignment(horizontal='center', vertical='center')

    ws6.row_dimensions[4].height = 24
    for ci, h in enumerate(['Status','Count'], 1):
        style_header(ws6.cell(4, ci), h, bg=MID_BLUE)
    ws6.column_dimensions['A'].width = 22
    ws6.column_dimensions['B'].width = 16

    status_df = summaries['by_status'].reset_index(drop=True)
    scols = {'Completed': GREEN, 'Pending': ORANGE, 'Cancelled': 'C0392B'}
    for ri, row in status_df.iterrows():
        r = ri + 5
        ws6.row_dimensions[r].height = 22
        color = scols.get(row.iloc[0], "888888")
        for ci, v in enumerate([row.iloc[0], row.iloc[1]], 1):
            c = ws6.cell(r, ci)
            c.value = v
            c.fill = PatternFill('solid', fgColor=color)
            c.font = Font(name='Calibri', size=11, bold=True, color=WHITE)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border()

    # Pie chart
    pie = PieChart()
    pie.title = "Order Status Distribution"
    pie.style = 10
    pie.width = 14
    pie.height = 10
    n6 = len(status_df)
    d = Reference(ws6, min_col=2, min_row=4, max_row=4+n6)
    pie.add_data(d, titles_from_data=True)
    labels6 = Reference(ws6, min_col=1, min_row=5, max_row=4+n6)
    pie.set_categories(labels6)
    ws6.add_chart(pie, f"A{n6+8}")

    wb.save(out_path)
    return out_path

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    INPUT  = "/home/claude/data_automation/raw_sales_data.csv"
    OUTPUT = "/home/claude/data_automation/Sales_Report_Cleaned.xlsx"

    print("📂 Loading raw data...")
    df_raw = load_data(INPUT)
    print(f"   {df_raw.shape[0]} rows x {df_raw.shape[1]} cols loaded")

    print("\n🧹 Running cleaning pipeline...")
    df_clean, cleaning_log = clean_data(df_raw)
    for entry in cleaning_log:
        print(f"   ✔ {entry['step']}: {entry['fixed']}")

    print("\n📊 Computing summaries...")
    summaries = compute_summaries(df_clean)

    print("\n📝 Writing Excel report...")
    out = write_excel_report(df_clean, cleaning_log, summaries, OUTPUT)
    print(f"\n✅ Report saved → {out}")

    log_path = "/home/claude/data_automation/cleaning_log.json"
    with open(log_path, 'w') as f:
        json.dump(cleaning_log, f, indent=2, default=str)
    print(f"📋 Cleaning log → {log_path}")
